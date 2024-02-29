import openpyxl

#to read from excel
book = openpyxl.load_workbook("/Users/mohammad.faiz/Downloads/file_to_file_input.xlsx")
sheet = book.active #will work on the active sheet
cell = sheet.cell(row=1, column=2) #extract text from given excel
print(cell.value)

#write in excel
sheet.cell(row=3, column=1).value="Faiz"
print(sheet.cell(row=3, column=1).value)


#To know number of rpw and column sheet level
print(sheet.max_row)
print(sheet.max_column)

#extract value A2
print(sheet["A2"].value)

#print all value of sheet
for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value)